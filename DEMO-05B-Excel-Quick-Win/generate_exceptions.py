"""Generate KN-Delivery-Exceptions-500.xlsx — 500 unstructured delivery exception notes.

The point of the Excel quick-win demo: 500 rows of free-text notes that, without Copilot,
would take a service-desk lead hours to read, classify, and prioritise. With one prompt
in Excel + Copilot, it becomes structured action in under a minute.
"""

import random
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

random.seed(20260525)

CUSTOMERS = [
    "Contoso Industries", "Fabrikam Pharma", "Northwind Traders", "Adventure Works",
    "Tailspin Toys", "Wide World Importers", "Litware Chemicals", "Proseware Electronics",
]
MODES = ["Sea", "Air", "Road"]
ORIGINS = ["Shanghai", "Singapore", "Rotterdam", "Hamburg", "Antwerp", "Los Angeles",
           "Frankfurt", "Dubai", "Hong Kong", "Mumbai", "Felixstowe", "Genoa"]
DESTS = ["Hamburg", "Rotterdam", "New York", "Chicago", "Frankfurt", "Berlin",
         "London", "Madrid", "Milan", "Stockholm", "Warsaw", "Dublin"]

# Realistic exception note templates. Mixed casing/punctuation on purpose (real inbox texts).
TEMPLATES = [
    # Customs
    "Customs hold at {dest} - missing commercial invoice. AM notified, waiting on shipper.",
    "HS code mismatch flagged by customs in {dest}. Documents being amended, ETA slip ~48h.",
    "shipment stuck at customs {dest} since {days} days - dual-use export licence pending",
    "Customs requested additional certificate of origin - shipper unresponsive for 2 days.",
    # Capacity / port
    "Port congestion at {origin} - vessel waiting at anchorage, 6 day delay so far.",
    "carrier rolled the booking at {origin}, next vessel only Friday. Customer notified.",
    "no space confirmed for week 12 sailing ex {origin}. Need to reroute via {alt}.",
    "Truck capacity tight {origin}->{dest}. Sub-contractor declined, looking for alternative.",
    # Damage / handling
    "2 pallets damaged on arrival {dest} - photos taken, claim being prepared by ops.",
    "DAMAGED CRATE reported by consignee. Value approx EUR 45k. Insurance file opened.",
    "Cold-chain breach detected en route to {dest} - temperature excursion 4h above +8C.",
    "carton count short by 3 at delivery {dest}. POD signed under protest. Investigating.",
    # Documentation
    "B/L not yet released by shipper - payment dispute upstream. Cargo at risk of demurrage.",
    "EUR1 form missing for {dest} shipment - customer cannot claim preferential duty.",
    "wrong consignee name on AWB - amendment fee USD 50, customer asked to cover.",
    "Packing list does not match invoice qty (off by 2 units). Awaiting clarification.",
    # Weather / external
    "Storm closure at {origin} port - all outbound delayed by 2 days minimum.",
    "Strike action at {dest} terminal - cargo cannot be released until Monday.",
    "Wildfire smoke shut down {dest} airport for 18h - air shipments rebooked.",
    # Internal / service
    "wrong delivery address keyed by ops - re-delivery scheduled tomorrow, fee waived",
    "Driver no-show for collection {origin}. Backup arranged, 4h delay.",
    "Forklift damaged a unit during transhipment in {origin}. Internal write-off.",
    # Routine / OK-ish
    "Shipment arrived on time, no issues - customer confirmed receipt.",
    "ETA pushed back 1 day due to vessel schedule revision. No customer impact.",
    "Late despatch from shipper warehouse by ~6h, recovered with expedited road leg.",
    # Invoicing / billing
    "Demurrage being charged - 4 days storage at {dest} terminal, claim against shipper.",
    "Customer disputes detention charge USD 800 - requesting waiver due to customs hold.",
    "Surcharge applied for hazmat handling - customer not informed at booking, complaint open.",
]

ALT_PORTS = ["Antwerp", "Le Havre", "Valencia", "Trieste"]

start_date = datetime(2026, 1, 6)

rows = []
for i in range(500):
    sid = f"KN-{random.randint(100000, 999999)}"
    cust = random.choice(CUSTOMERS)
    mode = random.choices(MODES, weights=[5, 3, 2])[0]
    origin = random.choice(ORIGINS)
    dest = random.choice([d for d in DESTS if d != origin])
    lane = f"{origin} -> {dest}"
    note = random.choice(TEMPLATES).format(
        dest=dest, origin=origin, days=random.randint(2, 7), alt=random.choice(ALT_PORTS)
    )
    # Random date in Q1 2026
    date = start_date + timedelta(days=random.randint(0, 89))
    rows.append([date.strftime("%Y-%m-%d"), sid, cust, mode, lane, note])

wb = Workbook()
ws = wb.active
ws.title = "Exceptions"
headers = ["Date", "Shipment_ID", "Customer", "Mode", "Lane", "Exception_Note"]
ws.append(headers)
for r in rows:
    ws.append(r)

# Header style
hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="003369")
for c in ws[1]:
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="left", vertical="center")

# Column widths
widths = [12, 14, 22, 8, 30, 90]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w

# Freeze header
ws.freeze_panes = "A2"

# Briefing sheet
brief = wb.create_sheet("Briefing", 0)
brief["A1"] = "KN Delivery Exceptions — Q1 2026"
brief["A1"].font = Font(bold=True, size=16, color="003369")
lines = [
    "",
    "500 raw delivery exception notes captured by the K+N service desk during Q1 2026.",
    "Each row is one shipment that triggered an exception flag - the Note column is free text",
    "(real-world style: mixed casing, abbreviations, internal jargon).",
    "",
    "Without Copilot: a service-desk lead would read all 500 notes, mentally categorise them,",
    "tally severity, then build a summary - typically a half-day of work.",
    "",
    "With Copilot in Excel: one prompt - 30 seconds - structured categories, severity, and",
    "recommended action per row, plus an executive summary sheet.",
    "",
    "Try the hero prompt in the demo briefing (DEMO-05B).",
]
for i, line in enumerate(lines, 2):
    brief.cell(row=i, column=1, value=line)
brief.column_dimensions["A"].width = 100

out = Path(__file__).parent / "KN-Delivery-Exceptions-500.xlsx"
wb.save(out)
print(f"Wrote {out} ({len(rows)} rows)")
