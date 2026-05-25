"""
Generate a realistic, "messy enough to be impressive" KN-style shipments
dataset for the Copilot in Excel demo.

Output: KN-Q1-2026-Shipments.xlsx in the same folder.

Run:  py _generate_shipments_xlsx.py
"""

from __future__ import annotations

import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

random.seed(20260524)

OUT = Path(__file__).with_name("KN-Q1-2026-Shipments.xlsx")

# --- Reference data ----------------------------------------------------------

CUSTOMERS = [
    # (Name, segment, default mode weight, base margin bias)
    ("Contoso Industries", "Industrial", {"Sea": 0.55, "Air": 0.30, "Road": 0.15}, -0.04),
    ("Fabrikam Pharma", "Pharma", {"Sea": 0.10, "Air": 0.75, "Road": 0.15}, +0.08),
    ("Northwind Traders", "Retail", {"Sea": 0.70, "Air": 0.10, "Road": 0.20}, +0.05),
    ("Adventure Works", "Automotive", {"Sea": 0.45, "Air": 0.10, "Road": 0.45}, +0.02),
    ("Tailspin Toys", "Consumer Goods", {"Sea": 0.80, "Air": 0.05, "Road": 0.15}, +0.06),
    ("Wide World Importers", "Food & Beverage", {"Sea": 0.60, "Air": 0.05, "Road": 0.35}, -0.01),
    ("Litware Chemicals", "Chemicals", {"Sea": 0.65, "Air": 0.10, "Road": 0.25}, +0.03),
    ("Proseware Electronics", "High-Tech", {"Sea": 0.30, "Air": 0.55, "Road": 0.15}, +0.07),
]

SEA_LANES = [
    ("Shanghai", "Hamburg"),
    ("Shenzhen", "Rotterdam"),
    ("Ningbo", "Hamburg"),
    ("Singapore", "Antwerp"),
    ("Busan", "Hamburg"),
    ("Tokyo", "Los Angeles"),
    ("Mumbai", "Genoa"),
    ("Santos", "Hamburg"),
]
AIR_LANES = [
    ("Frankfurt", "Chicago"),
    ("Hong Kong", "Frankfurt"),
    ("Shanghai", "Luxembourg"),
    ("Mumbai", "Brussels"),
    ("Seoul", "Amsterdam"),
    ("Dubai", "Frankfurt"),
    ("Atlanta", "Frankfurt"),
    ("Singapore", "Frankfurt"),
]
ROAD_LANES = [
    ("Hamburg", "Berlin"),
    ("Hamburg", "Munich"),
    ("Rotterdam", "Hamburg"),
    ("Antwerp", "Frankfurt"),
    ("Hamburg", "Warsaw"),
    ("Duisburg", "Lyon"),
    ("Hamburg", "Prague"),
    ("Bremerhaven", "Stuttgart"),
]

SERVICES = {
    "Sea": ["FCL Standard", "FCL Premium", "LCL Consolidated", "Reefer"],
    "Air": ["Air Express", "Air Standard", "Air Pharma Temp", "Charter"],
    "Road": ["FTL", "LTL", "Express Courier", "Tail-lift"],
}

CONTAINER_TYPES = {
    "Sea": ["20'DC", "40'DC", "40'HC", "40'RF", "LCL"],
    "Air": ["ULD AKE", "ULD AMA", "Loose", "Pharma ULD"],
    "Road": ["13.6m Curtainsider", "13.6m Box", "Refrigerated", "Tanker"],
}

NOTES_POOL_GOOD = [
    "Delivered on time",
    "POD received",
    "Customer confirmed receipt",
    "No issues reported",
    "Cleared customs same day",
    "",
    "",
    "",
]
NOTES_POOL_WATCH = [
    "Minor delay at origin",
    "Port congestion noted",
    "Documentation rework required",
    "Customs hold 1 day",
    "Driver swap at hub",
]
NOTES_POOL_BAD = [
    "Demurrage charges accrued - awaiting customer dispute",
    "Temperature excursion logged - investigation open",
    "Damaged seal - claim filed",
    "Missed cut-off, rebooked",
    "Reefer alarm - product release pending QA",
    "Detention exceeded SLA - cost overrun",
]


def pick_mode(weights: dict[str, float]) -> str:
    r = random.random()
    cum = 0.0
    for m, w in weights.items():
        cum += w
        if r <= cum:
            return m
    return "Sea"


def gen_shipment(i: int) -> dict:
    cust_name, segment, mode_w, margin_bias = random.choice(CUSTOMERS)
    mode = pick_mode(mode_w)

    if mode == "Sea":
        origin, dest = random.choice(SEA_LANES)
        transit = random.randint(22, 42)
        revenue = round(random.uniform(2_200, 24_500), 2)
        weight = random.choice([18_000, 21_000, 24_000, 26_000, 28_000])
    elif mode == "Air":
        origin, dest = random.choice(AIR_LANES)
        transit = random.randint(2, 6)
        revenue = round(random.uniform(4_800, 58_000), 2)
        weight = random.choice([1_200, 2_400, 4_800, 7_500, 12_000])
    else:
        origin, dest = random.choice(ROAD_LANES)
        transit = random.randint(1, 5)
        revenue = round(random.uniform(480, 4_200), 2)
        weight = random.choice([3_500, 12_000, 18_000, 22_000])

    # Cost as fraction of revenue, biased per customer
    margin_pct = random.gauss(0.10 + margin_bias, 0.09)
    margin_pct = max(-0.25, min(0.40, margin_pct))
    cost = round(revenue * (1 - margin_pct), 2)
    margin = round(revenue - cost, 2)

    # Date in Q1 2026
    start = date(2026, 1, 1)
    ship_date = start + timedelta(days=random.randint(0, 89))

    # Status
    sla_breach = transit > {"Sea": 38, "Air": 5, "Road": 4}[mode]
    if margin_pct < -0.05 or sla_breach:
        status = random.choices(
            ["Delayed", "Exception", "Delivered"], weights=[0.4, 0.3, 0.3]
        )[0]
        notes = random.choice(NOTES_POOL_BAD)
    elif margin_pct < 0.05:
        status = random.choices(["Delivered", "Delayed"], weights=[0.7, 0.3])[0]
        notes = random.choice(NOTES_POOL_WATCH)
    else:
        status = "Delivered"
        notes = random.choice(NOTES_POOL_GOOD)

    # Intentional minor messiness: occasional mixed casing on origin/dest
    if random.random() < 0.08:
        origin = origin.upper()
    if random.random() < 0.05:
        dest = dest.lower()

    return {
        "Shipment_ID": f"KN-{2026}{i:05d}",
        "Date": ship_date,
        "Mode": mode,
        "Service": random.choice(SERVICES[mode]),
        "Origin": origin,
        "Destination": dest,
        "Customer": cust_name,
        "Segment": segment,
        "Container_Type": random.choice(CONTAINER_TYPES[mode]),
        "Weight_kg": weight,
        "Transit_Days": transit,
        "Revenue_EUR": revenue,
        "Cost_EUR": cost,
        "Margin_EUR": margin,
        "Status": status,
        "Notes": notes,
    }


def main() -> None:
    rows = [gen_shipment(i) for i in range(1, 1501)]

    wb = Workbook()

    # --- Sheet 1: Raw Shipments ---------------------------------------------
    ws = wb.active
    ws.title = "Raw Shipments"

    headers = list(rows[0].keys())
    ws.append(headers)

    for h_cell in ws[1]:
        h_cell.font = Font(bold=True, color="FFFFFF")
        h_cell.fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        h_cell.alignment = Alignment(horizontal="left", vertical="center")

    for r in rows:
        ws.append([r[h] for h in headers])

    # Format currency columns
    for col_name in ("Revenue_EUR", "Cost_EUR", "Margin_EUR"):
        col_idx = headers.index(col_name) + 1
        letter = get_column_letter(col_idx)
        for cell in ws[letter][1:]:
            cell.number_format = '#,##0.00 "\u20ac"'

    # Format date column
    date_col = get_column_letter(headers.index("Date") + 1)
    for cell in ws[date_col][1:]:
        cell.number_format = "yyyy-mm-dd"

    # Column widths
    widths = {
        "Shipment_ID": 14, "Date": 12, "Mode": 8, "Service": 18,
        "Origin": 16, "Destination": 16, "Customer": 22, "Segment": 16,
        "Container_Type": 18, "Weight_kg": 11, "Transit_Days": 13,
        "Revenue_EUR": 14, "Cost_EUR": 14, "Margin_EUR": 14,
        "Status": 12, "Notes": 55,
    }
    for h, w in widths.items():
        ws.column_dimensions[get_column_letter(headers.index(h) + 1)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # --- Sheet 2: Customers --------------------------------------------------
    ws2 = wb.create_sheet("Customers")
    ws2.append(["Customer", "Segment", "Account Manager", "Region", "Tier"])
    for h_cell in ws2[1]:
        h_cell.font = Font(bold=True, color="FFFFFF")
        h_cell.fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
    tier_map = {
        "Contoso Industries": ("Anna Becker", "EMEA", "Strategic"),
        "Fabrikam Pharma": ("Julia Werner", "EMEA", "Strategic"),
        "Northwind Traders": ("Marco Lentz", "APAC", "Key"),
        "Adventure Works": ("Sven Holm", "EMEA", "Key"),
        "Tailspin Toys": ("Priya Raman", "APAC", "Standard"),
        "Wide World Importers": ("Anna Becker", "Americas", "Key"),
        "Litware Chemicals": ("Julia Werner", "EMEA", "Standard"),
        "Proseware Electronics": ("Sven Holm", "APAC", "Strategic"),
    }
    for name, segment, _, _ in CUSTOMERS:
        am, region, tier = tier_map[name]
        ws2.append([name, segment, am, region, tier])
    for col, w in zip("ABCDE", (24, 18, 18, 12, 12)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    # --- Sheet 3: Q4 2025 Baseline ------------------------------------------
    ws3 = wb.create_sheet("Q4 2025 Baseline")
    ws3.append(["Customer", "Mode", "Shipments_Q4", "Revenue_Q4_EUR", "Margin_Q4_EUR", "Margin_Q4_Pct"])
    for h_cell in ws3[1]:
        h_cell.font = Font(bold=True, color="FFFFFF")
        h_cell.fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")

    for cust_name, _, mode_w, margin_bias in CUSTOMERS:
        for mode in ("Sea", "Air", "Road"):
            if mode_w.get(mode, 0) < 0.05:
                continue
            shipments_q4 = int(60 * mode_w[mode] * random.uniform(0.8, 1.3))
            rev_avg = {"Sea": 12_000, "Air": 28_000, "Road": 2_100}[mode]
            revenue_q4 = round(shipments_q4 * rev_avg * random.uniform(0.85, 1.2), 2)
            margin_pct_q4 = round(0.11 + margin_bias + random.uniform(-0.03, 0.03), 4)
            margin_q4 = round(revenue_q4 * margin_pct_q4, 2)
            ws3.append([cust_name, mode, shipments_q4, revenue_q4, margin_q4, margin_pct_q4])

    for c_letter in ("D", "E"):
        for cell in ws3[c_letter][1:]:
            cell.number_format = '#,##0.00 "\u20ac"'
    for cell in ws3["F"][1:]:
        cell.number_format = "0.00%"
    for col, w in zip("ABCDEF", (24, 8, 14, 18, 18, 14)):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A2"

    # --- Sheet 4: Briefing ---------------------------------------------------
    ws4 = wb.create_sheet("Briefing", 0)
    ws4.append(["KN Q1 2026 Shipments - Margin Analysis Workbook"])
    ws4["A1"].font = Font(bold=True, size=16, color="107C41")
    ws4.append([])
    ws4.append([
        "1,500 shipments across Sea, Air and Road for Q1 2026. "
        "Customers include strategic and key accounts. "
        "Goal: identify loss-making lanes, customers and modes - and forecast Q2."
    ])
    ws4.append([])
    ws4.append(["Sheets:"])
    ws4.append(["  Raw Shipments", "Per-shipment data (1,500 rows)"])
    ws4.append(["  Customers", "Account directory with tier and AM"])
    ws4.append(["  Q4 2025 Baseline", "Prior-quarter benchmark for comparison"])
    ws4.append([])
    ws4.append(["Suggested workflow:"])
    ws4.append(["  1.", "Open Copilot in Excel - Tools menu - Edit with Copilot"])
    ws4.append(["  2.", "Ask Copilot to build a margin dashboard (PivotTables + charts)"])
    ws4.append(["  3.", "Ask Copilot for a Q2 outlook using FORECAST.LINEAR + a what-if scenario"])
    ws4.append(["  4.", "Optional (Frontier / Insider only): use =COPILOT() to tag each row with an AI risk flag"])
    ws4.append([])
    ws4.append(["Licensing:"])
    ws4.append(["  Edit with Copilot", "Requires M365 Copilot add-on (Win / Web / Mac)"])
    ws4.append(["  Python advanced analysis", "Requires M365 Copilot add-on (Win / Web)"])
    ws4.append(["  =COPILOT() function", "Frontier Program + M365 Insider only"])
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 70
    ws4.sheet_view.showGridLines = False

    wb.save(OUT)
    print(f"Wrote {OUT} ({OUT.stat().st_size / 1024:.1f} KB, {len(rows)} shipment rows)")


if __name__ == "__main__":
    main()
